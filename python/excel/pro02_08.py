# 表单

import openpyxl

book = openpyxl.load_workbook('sheets.xlsx')
print(book.sheetnames)
sheet_active = book.active
print(type(sheet_active))
sheet = book["Sheet1"]
print(sheet.title)
# 创建一个工作表 指定位置，指定名称
# book.create_sheet("Sheet1", 0)
# book.create_sheet("Pete", 2)
print(book.sheetnames)
# # 删除一个工作表
# book.remove_sheet(sheet)
# print(book.sheetnames)

# 改变工作簿的颜色
sheet.sheet_properties.tabColor = "0072BA"
book.save('sheets.xlsx')
