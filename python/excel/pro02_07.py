# openpyxl 维度
import random

import openpyxl

book = openpyxl.Workbook()
sheet = book.active
sheet["A3"] = "Pete"
sheet["C3"] = "Janne"
rows = [(random.randint(23, 89) for _ in range(3)),
        (random.randint(23, 89) for _ in range(3)),
        (random.randint(23, 89) for _ in range(3)),
        (random.randint(23, 89) for _ in range(3)),
        (random.randint(23, 89) for _ in range(3)),
        (random.randint(23, 89) for _ in range(3)),
        (random.randint(23, 89) for _ in range(3)),
        (random.randint(23, 89) for _ in range(3)),
        (random.randint(23, 89) for _ in range(3))
        ]
for row in rows:
    sheet.append(row)
    pass

print("dimensions: {0}".format(sheet.dimensions))
print(type(sheet.dimensions))
print("Min row: {0}".format(sheet.min_row))  # 返回包含数据的最小行索引，索引从1开始
print("Max row: {0}".format(sheet.max_row))  # 返回包含数据的最大行索引，索引从1开始
print("Min col: {0}".format(sheet.min_column))
print("Max col: {0}".format(sheet.max_column))

# for row in sheet.rows:
#     for cell in row:
#         print(cell.value, end=" ")
#         pass
#
#     pass

# 打印表单的数据 ，与上一个方法进行对比
for c1, c2, c3 in sheet[sheet.dimensions]:
    print(c1.value, c2.value, c3.value)
    pass

book.save("pro_02_07.xlsx")
