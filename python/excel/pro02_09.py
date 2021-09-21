# 合并单元格
# 导入 Alignment类设置格式

from openpyxl import Workbook
from openpyxl.styles import Alignment

book = Workbook()
sheet = book.active
# 合并单元格
sheet.merge_cells("A1:B2")
# 设置单元格
cell = sheet.cell(row=1, column=1)
# 单元格取值
cell.value = "Pete"
# 设置字体位置
cell.alignment = Alignment(horizontal="center", vertical="center")
book.save("pro02_09.xlsx")

