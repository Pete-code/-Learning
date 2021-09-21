# openpyxl 按行遍历迭代
import random

import openpyxl
# 导入类
book = openpyxl.Workbook()
# 创建一个工作薄
sheet = book.active
# 使用随机数生成列表
rows = ((random.randint(45, 89) for _ in range(3)),
        (random.randint(45, 89) for _ in range(3)),
        (random.randint(45, 89) for _ in range(3)),
        (random.randint(45, 89) for _ in range(3)),
        (random.randint(45, 89) for _ in range(3))
        )

# 遍历列表，将列表添加到工作薄中
for row in rows:
    sheet.append(row)
    pass

# 遍历行读取内容，按行迭代
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=5, max_col=3): # 划定遍历的范围，获取每一行的元组
    print(type(row))
    for cell in row:  # 获取每一个单元格中的数据
        print(cell.value, end=" ")
        pass
    pass
# 按列迭代
for col  in sheet.iter_cols(min_row=1, min_col=1, max_row=5, max_col=3): # 划定遍历的范围，获取每一行的元组
    print(type(col))
    for cell in col:  # 获取每一个单元格中的数据
        print(cell.value, end=" ")
        pass
    pass

# 记住保存创建的文件
book.save("pro02_04.xlsx")

